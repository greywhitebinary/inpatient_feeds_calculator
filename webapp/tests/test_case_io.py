import os
import sys
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from case_io import export_case_record_workbook, import_case_record_workbook
from data import load_master_formulas, load_master_modulars, load_master_ons


class CaseRecordTests(unittest.TestCase):
    def test_default_website_metadata_is_safe_and_not_hyperlinked(self):
        payload = export_case_record_workbook(
            {}, load_master_formulas().head(0), load_master_modulars().head(0)
        )
        workbook = load_workbook(BytesIO(payload))
        sheet = workbook["Case record"]

        self.assertEqual(sheet["A2"].value, "Calculator website")
        self.assertEqual(sheet["B2"].value, "To be added after deployment")
        self.assertIsNone(sheet["B2"].hyperlink)
        self.assertIn("does not retain case records", sheet["A4"].value)
        self.assertIn("hosted session processes entered values", sheet["A4"].value)

    def test_live_configured_website_is_hyperlinked(self):
        with patch.dict(
            os.environ,
            {"CALCULATOR_WEBSITE_URL": "https://feeds.example.org/calculator"},
        ):
            payload = export_case_record_workbook(
                {}, load_master_formulas().head(0), load_master_modulars().head(0)
            )
        sheet = load_workbook(BytesIO(payload))["Case record"]

        self.assertEqual(sheet["B2"].value, "https://feeds.example.org/calculator")
        self.assertEqual(
            sheet["B2"].hyperlink.target, "https://feeds.example.org/calculator"
        )

    def test_localhost_configuration_is_not_written_to_workbook(self):
        with patch.dict(
            os.environ, {"CALCULATOR_WEBSITE_URL": "http://localhost:8501"}
        ):
            payload = export_case_record_workbook(
                {}, load_master_formulas().head(0), load_master_modulars().head(0)
            )
        sheet = load_workbook(BytesIO(payload))["Case record"]

        self.assertEqual(sheet["B2"].value, "To be added after deployment")
        self.assertIsNone(sheet["B2"].hyperlink)

    def test_import_accepts_workbook_without_new_website_field(self):
        formulas = load_master_formulas().head(0)
        modulars = load_master_modulars().head(0)
        payload = export_case_record_workbook(
            {"case_record_label": "Older record"}, formulas, modulars
        )
        workbook = load_workbook(BytesIO(payload))
        workbook["Case record"].delete_rows(2, 1)
        legacy = BytesIO()
        workbook.save(legacy)

        restored, _, _, _ = import_case_record_workbook(BytesIO(legacy.getvalue()))

        self.assertEqual(restored["case_record_label"], "Older record")

    def test_import_accepts_case_record_without_my_ons_sheet(self):
        payload = export_case_record_workbook(
            {"case_record_label": "Pre-ONS record"},
            load_master_formulas().head(0),
            load_master_modulars().head(0),
        )
        workbook = load_workbook(BytesIO(payload))
        del workbook["My ONS"]
        legacy = BytesIO()
        workbook.save(legacy)

        restored, _, _, restored_ons = import_case_record_workbook(
            BytesIO(legacy.getvalue())
        )

        self.assertEqual(restored["case_record_label"], "Pre-ONS record")
        self.assertTrue(restored_ons.empty)

    def test_import_accepts_legacy_plan_goal_mirrors(self):
        payload = export_case_record_workbook(
            {"case_record_label": "Older goal record"},
            load_master_formulas().head(0),
            load_master_modulars().head(0),
        )
        workbook = load_workbook(BytesIO(payload))
        inputs = workbook["Case inputs"]
        inputs.append(["icu_total_energy_target", "1750.0"])
        legacy = BytesIO()
        workbook.save(legacy)

        restored, _, _, _ = import_case_record_workbook(BytesIO(legacy.getvalue()))

        self.assertEqual(restored["icu_total_energy_target"], 1750.0)

    def test_round_trip_includes_inputs_and_formulary_snapshot(self):
        formulas = load_master_formulas().head(1)
        modulars = load_master_modulars().head(1)
        state = {
            "case_record_label": "Ward nutrition review",
            "assessment_age": 67.0,
            "assessment_sex": "Female",
            "assessment_energy_target": 1750.0,
            "feed_candidates": [formulas.iloc[0]["name"]],
            "chosen_modulars": [modulars.iloc[0]["name"]],
            "en_hydration_schedule_format": "qXh",
            "en_hydration_interval_hours": 4,
            f"modular_units_{modulars.iloc[0]['id']}": 2.0,
        }

        payload = export_case_record_workbook(state, formulas, modulars)
        restored, restored_formulas, restored_modulars, restored_ons = (
            import_case_record_workbook(BytesIO(payload))
        )

        self.assertEqual(restored["assessment_age"], 67.0)
        self.assertEqual(restored["feed_candidates"], [formulas.iloc[0]["name"]])
        self.assertEqual(restored["en_hydration_schedule_format"], "qXh")
        self.assertEqual(restored["en_hydration_interval_hours"], 4)
        self.assertEqual(restored_formulas["name"].tolist(), formulas["name"].tolist())
        self.assertEqual(restored_modulars["id"].tolist(), modulars["id"].tolist())
        self.assertTrue(restored_ons.empty)

    def test_round_trip_preserves_ons_snapshot_and_order(self):
        formulas = load_master_formulas().head(1)
        modulars = load_master_modulars().head(0)
        ons = (
            load_master_ons()
            .loc[lambda frame: frame["name"] == "BOOST Plus Calories — Vanilla"]
            .copy()
        )
        product_id = ons.iloc[0]["id"]
        state = {
            "scenario_standard_chosen_ons": [ons.iloc[0]["name"]],
            f"scenario_standard_ons_containers_{product_id}": 1.0,
            f"scenario_standard_ons_times_{product_id}": 2.0,
        }

        payload = export_case_record_workbook(state, formulas, modulars, ons)
        restored, _, _, restored_ons = import_case_record_workbook(BytesIO(payload))

        self.assertEqual(
            restored["scenario_standard_chosen_ons"],
            ["BOOST Plus Calories — Vanilla"],
        )
        self.assertEqual(
            restored[f"scenario_standard_ons_containers_{product_id}"], 1.0
        )
        self.assertEqual(restored[f"scenario_standard_ons_times_{product_id}"], 2.0)
        self.assertEqual(
            restored_ons["name"].tolist(), ["BOOST Plus Calories — Vanilla"]
        )

    def test_round_trip_preserves_an_unentered_value(self):
        formulas = load_master_formulas().head(0)
        modulars = load_master_modulars().head(0)
        payload = export_case_record_workbook(
            {"case_record_label": "Blank example", "assessment_energy_target": None},
            formulas,
            modulars,
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        self.assertIn("assessment_energy_target", restored)
        self.assertIsNone(restored["assessment_energy_target"])

    def test_round_trip_preserves_an_ordered_flush_schedule(self):
        state = {
            "case_record_label": "Running flush order",
            "scenario_standard_regimen_source": "Reviewing a feed already running",
            "scenario_standard_hydration_entry_mode": "Enter flushes as ordered",
            "scenario_standard_peri_feed_flush_pattern": "Before and after each feed",
            "scenario_standard_peri_feed_flush_volume_ml": 150.0,
            "scenario_standard_ordered_flush_times_per_day": 1,
            "scenario_standard_ordered_flush_volume_ml": 150.0,
        }
        payload = export_case_record_workbook(
            state, load_master_formulas().head(0), load_master_modulars().head(0)
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        for key, value in state.items():
            self.assertEqual(restored[key], value)

    def test_round_trip_preserves_a_rate_and_duration_order(self):
        state = {
            "case_record_label": "Bolus by rate",
            "scenario_standard_schedule_type": "Intermittent",
            "scenario_standard_order_entry_form": "A rate in mL/hour, run for a set time each feed",
            "scenario_standard_ordered_entry_form": "A rate in mL/hour, run for a set time each feed",
            "scenario_standard_hours_per_feed": 2.0,
            "scenario_standard_feeds_per_day": 3,
            "scenario_standard_ordered_rate_ml_hr": 180.0,
        }
        payload = export_case_record_workbook(
            state, load_master_formulas().head(0), load_master_modulars().head(0)
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        for key, value in state.items():
            self.assertEqual(restored[key], value)

    def test_round_trip_preserves_a_reviewed_running_order(self):
        # A record saved while reviewing carries the running-shape field. It was
        # unregistered at first, so saving in that mode produced a file that
        # would not reopen.
        state = {
            "case_record_label": "Reviewed on admission",
            "scenario_standard_regimen_source": "Reviewing a feed already running",
            "scenario_standard_running_shape": (
                "Intermittent, each feed run at a rate for a set time"
            ),
            "scenario_standard_hours_per_feed": 2.0,
            "scenario_standard_feeds_per_day": 3,
        }
        payload = export_case_record_workbook(
            state, load_master_formulas().head(0), load_master_modulars().head(0)
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        for key, value in state.items():
            self.assertEqual(restored[key], value)

    def test_a_record_saved_after_visiting_the_propofol_page_reopens(self):
        # Conditional rates are keyed by the sedation condition they belong to.
        # Only two ids were registered by name, so the single-rate page's own
        # condition produced a field the importer rejected, and any record
        # saved after visiting that page could not be reopened.
        state = {
            "case_record_label": "Seen on the Propofol page",
            "scenario_propofol_conditional_projected_rate_ml_hr": 45.0,
            "scenario_propofol_conditional_projected_rate_user_edited": True,
            "scenario_propofol_conditional_lower_rate_ml_hr": 30.0,
        }
        payload = export_case_record_workbook(
            state, load_master_formulas().head(0), load_master_modulars().head(0)
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        for key, value in state.items():
            self.assertEqual(restored[key], value)

    def test_round_trip_preserves_a_daily_total_order(self):
        state = {
            "case_record_label": "Written as a daily total",
            "scenario_standard_order_entry_form": "A total volume in mL per day",
            "scenario_standard_ordered_daily_volume_ml": 1080.0,
        }
        payload = export_case_record_workbook(
            state, load_master_formulas().head(0), load_master_modulars().head(0)
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        for key, value in state.items():
            self.assertEqual(restored[key], value)

    def test_import_rejects_an_invalid_order_entry_form(self):
        payload = export_case_record_workbook(
            {"scenario_standard_order_entry_form": "A rate in mL/hour"},
            load_master_formulas().head(0),
            load_master_modulars().head(0),
        )
        workbook = load_workbook(BytesIO(payload))
        workbook["Case inputs"]["B2"] = '"Any old way"'
        edited = BytesIO()
        workbook.save(edited)

        with self.assertRaisesRegex(ValueError, "unsupported order entry form"):
            import_case_record_workbook(BytesIO(edited.getvalue()))

    def test_record_saved_before_the_new_flush_fields_still_opens(self):
        # A record written before this work simply lacks the new fields. It must
        # open unchanged, with the defaults supplied at seeding time rather than
        # by any conversion step here.
        payload = export_case_record_workbook(
            {
                "case_record_label": "Pre-change record",
                "assessment_age": 67,
                "scenario_standard_schedule_type": "Continuous / cyclic",
                "scenario_standard_feeding_hours": 23.0,
                "scenario_standard_hydration_flushes": 6,
            },
            load_master_formulas().head(0),
            load_master_modulars().head(0),
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        self.assertEqual(restored["scenario_standard_feeding_hours"], 23.0)
        self.assertEqual(restored["scenario_standard_hydration_flushes"], 6)
        self.assertNotIn("scenario_standard_regimen_source", restored)
        self.assertNotIn("scenario_standard_hydration_entry_mode", restored)

    def test_import_rejects_an_invalid_hydration_entry_mode(self):
        payload = export_case_record_workbook(
            {"scenario_standard_hydration_entry_mode": "Enter flushes as ordered"},
            load_master_formulas().head(0),
            load_master_modulars().head(0),
        )
        workbook = load_workbook(BytesIO(payload))
        workbook["Case inputs"]["B2"] = '"Guess the flushes"'
        edited = BytesIO()
        workbook.save(edited)

        with self.assertRaisesRegex(ValueError, "unsupported hydration entry mode"):
            import_case_record_workbook(BytesIO(edited.getvalue()))

    def test_download_excludes_generated_and_edited_chart_note_state(self):
        payload = export_case_record_workbook(
            {
                "case_record_label": "Chart-note exclusion check",
                "assessment_age": 67,
                "_chart_note_generated_en_plan": "Generated chart note",
                "_chart_note_editor_en_plan": "Clinician-edited chart note",
                "unrelated_widget_state": "Do not save",
            },
            load_master_formulas().head(1),
            load_master_modulars().head(1),
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        self.assertEqual(restored["case_record_label"], "Chart-note exclusion check")
        self.assertEqual(restored["assessment_age"], 67)
        self.assertNotIn("_chart_note_generated_en_plan", restored)
        self.assertNotIn("_chart_note_editor_en_plan", restored)
        self.assertNotIn("unrelated_widget_state", restored)

    def test_import_rejects_an_invalid_numeric_widget_value(self):
        payload = export_case_record_workbook(
            {"assessment_age": 67},
            load_master_formulas().head(0),
            load_master_modulars().head(0),
        )
        workbook = load_workbook(BytesIO(payload))
        inputs = workbook["Case inputs"]
        inputs["B2"] = '"sixty-seven"'
        edited = BytesIO()
        workbook.save(edited)

        with self.assertRaisesRegex(ValueError, "non-numeric.*assessment_age"):
            import_case_record_workbook(BytesIO(edited.getvalue()))

    def test_import_rejects_an_invalid_scenario_schedule(self):
        payload = export_case_record_workbook(
            {"scenario_standard_schedule_type": "Continuous / cyclic"},
            load_master_formulas().head(0),
            load_master_modulars().head(0),
        )
        workbook = load_workbook(BytesIO(payload))
        inputs = workbook["Case inputs"]
        inputs["B2"] = '"Unsupported schedule"'
        edited = BytesIO()
        workbook.save(edited)

        with self.assertRaisesRegex(ValueError, "unsupported schedule"):
            import_case_record_workbook(BytesIO(edited.getvalue()))

    def test_import_rejects_duplicate_case_fields(self):
        payload = export_case_record_workbook(
            {"assessment_age": 67},
            load_master_formulas().head(0),
            load_master_modulars().head(0),
        )
        workbook = load_workbook(BytesIO(payload))
        workbook["Case inputs"].append(["assessment_age", "68"])
        edited = BytesIO()
        workbook.save(edited)

        with self.assertRaisesRegex(ValueError, "duplicate field keys"):
            import_case_record_workbook(BytesIO(edited.getvalue()))

    def test_round_trip_preserves_both_en_scenarios(self):
        formulas = load_master_formulas().head(1)
        modulars = load_master_modulars().head(1)
        product_id = modulars.iloc[0]["id"]
        state = {
            "en_total_energy_target": 1800.0,
            "en_has_alternate_plan": True,
            "scenario_primary_include_propofol": False,
            "scenario_primary_propofol_rate": 0.0,
            "scenario_primary_selected_formula": formulas.iloc[0]["name"],
            f"scenario_primary_modular_doses_{product_id}": 2.0,
            "scenario_alternate_include_propofol": True,
            "scenario_alternate_propofol_rate": 20.0,
            "scenario_alternate_selected_formula": formulas.iloc[0]["name"],
            f"scenario_alternate_modular_doses_{product_id}": 4.0,
        }

        payload = export_case_record_workbook(state, formulas, modulars)
        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        self.assertFalse(restored["scenario_primary_include_propofol"])
        self.assertEqual(restored[f"scenario_primary_modular_doses_{product_id}"], 2.0)
        self.assertTrue(restored["scenario_alternate_include_propofol"])
        self.assertEqual(restored["scenario_alternate_propofol_rate"], 20.0)
        self.assertEqual(
            restored[f"scenario_alternate_modular_doses_{product_id}"], 4.0
        )

    def test_round_trip_preserves_independent_standard_and_icu_workflows(self):
        formulas = load_master_formulas().head(1)
        modulars = load_master_modulars().head(1)
        state = {
            "assessment_energy_target": 1800.0,
            "assessment_protein_target": 100.0,
            "assessment_water_target": 2000.0,
            "en_total_energy_target": 1800.0,
            "feed_candidates": [formulas.iloc[0]["name"]],
            "scenario_standard_feeding_hours": 18.0,
            "scenario_standard_ordered_rate_ml_hr": 55.0,
            "scenario_standard_ordered_formula_name": formulas.iloc[0]["name"],
            "scenario_standard_order_user_edited": True,
            "scenario_standard_order_reset_requested": True,
            "scenario_standard_hydration_schedule_format": "qXh",
            "scenario_standard_hydration_interval_hours": 4,
            "icu_total_energy_target": 1750.0,
            "icu_protein_target": 100.0,
            "icu_water_target": 2000.0,
            "icu_feed_candidates": [formulas.iloc[0]["name"]],
            "icu_planned_daily_intake_scenario": "higher",
            "scenario_lower_propofol_rate": 8.0,
            "scenario_lower_propofol_hours": 12.0,
            "scenario_lower_ordered_rate_ml_hr": 50.0,
            "scenario_higher_propofol_rate": 22.0,
            "scenario_higher_propofol_hours": 18.0,
            "scenario_higher_ordered_rate_ml_hr": 35.0,
            "assessment_activity_factor": 1.1,
            "assessment_stress_factor": 1.2,
        }

        payload = export_case_record_workbook(state, formulas, modulars)
        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        self.assertEqual(restored["scenario_standard_feeding_hours"], 18.0)
        self.assertEqual(restored["scenario_standard_ordered_rate_ml_hr"], 55.0)
        self.assertTrue(restored["scenario_standard_order_user_edited"])
        self.assertNotIn("scenario_standard_order_reset_requested", restored)
        self.assertEqual(restored["scenario_standard_hydration_schedule_format"], "qXh")
        self.assertEqual(restored["scenario_standard_hydration_interval_hours"], 4)
        self.assertEqual(restored["assessment_energy_target"], 1800.0)
        self.assertEqual(restored["assessment_protein_target"], 100.0)
        self.assertEqual(restored["assessment_water_target"], 2000.0)
        self.assertNotIn("en_total_energy_target", restored)
        self.assertNotIn("icu_total_energy_target", restored)
        self.assertEqual(restored["icu_planned_daily_intake_scenario"], "higher")
        self.assertEqual(restored["scenario_lower_propofol_rate"], 8.0)
        self.assertEqual(restored["scenario_lower_propofol_hours"], 12.0)
        self.assertEqual(restored["scenario_lower_ordered_rate_ml_hr"], 50.0)
        self.assertEqual(restored["scenario_higher_propofol_rate"], 22.0)
        self.assertEqual(restored["scenario_higher_propofol_hours"], 18.0)
        self.assertEqual(restored["scenario_higher_ordered_rate_ml_hr"], 35.0)
        self.assertEqual(restored["assessment_activity_factor"], 1.1)
        self.assertEqual(restored["assessment_stress_factor"], 1.2)

    def test_round_trip_preserves_the_shared_conditional_propofol_plan(self):
        formulas = load_master_formulas().head(1)
        modulars = load_master_modulars().head(1)
        product_id = modulars.iloc[0]["id"]
        state = {
            "assessment_energy_target": 1800.0,
            "icu_feed_candidates": [formulas.iloc[0]["name"]],
            "scenario_propofol_propofol_method": "Conditional EN rates",
            "scenario_propofol_prescription_target_pct": 110.0,
            "scenario_propofol_prescription_interruption_note": True,
            "scenario_propofol_lower_propofol_rate": 0.0,
            "scenario_propofol_higher_propofol_rate": 20.0,
            "scenario_propofol_higher_propofol_hours": 6.0,
            "scenario_propofol_selected_formula": formulas.iloc[0]["name"],
            "scenario_propofol_feeding_hours": 23.0,
            "scenario_propofol_conditional_lower_rate_ml_hr": 55.0,
            "scenario_propofol_conditional_higher_rate_ml_hr": 40.0,
            "scenario_propofol_conditional_lower_rate_user_edited": True,
            f"scenario_propofol_modular_doses_{product_id}": 3.0,
            "icu_planned_daily_intake_scenario": "higher",
            "scenario_higher_propofol_rate": 30.0,
        }

        payload = export_case_record_workbook(state, formulas, modulars)
        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        self.assertEqual(
            restored["scenario_propofol_propofol_method"], "Changing Propofol rates"
        )
        self.assertEqual(restored["scenario_propofol_prescription_target_pct"], 110.0)
        self.assertTrue(restored["scenario_propofol_prescription_interruption_note"])
        self.assertEqual(restored["scenario_propofol_higher_propofol_hours"], 6.0)
        self.assertEqual(
            restored["scenario_propofol_conditional_lower_rate_ml_hr"], 55.0
        )
        self.assertTrue(
            restored["scenario_propofol_conditional_lower_rate_user_edited"]
        )
        self.assertEqual(restored[f"scenario_propofol_modular_doses_{product_id}"], 3.0)
        self.assertNotIn("icu_planned_daily_intake_scenario", restored)
        self.assertNotIn("scenario_higher_propofol_rate", restored)

    def test_old_height_fields_migrate_to_centimetres_on_import(self):
        formulas = load_master_formulas().head(0)
        modulars = load_master_modulars().head(0)
        payload = export_case_record_workbook(
            {
                "assessment_height_unit": "m",
                "assessment_height_m": 1.65,
            },
            formulas,
            modulars,
        )

        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))

        self.assertEqual(restored["assessment_height_cm"], 165.0)
        self.assertNotIn("assessment_height_unit", restored)
        self.assertNotIn("assessment_height_m", restored)


class IVFluidRecordTests(unittest.TestCase):
    """An entered IV must survive save and reload, and a bad one must be refused."""

    def test_iv_entries_round_trip(self):
        state = {
            "case_record_label": "IV case",
            "assessment_iv_fluid_0": "D5 1/2 NS",
            "assessment_iv_rate_0": 100.0,
            "assessment_iv_fluid_1": "",
            "assessment_iv_rate_1": None,
        }
        payload = export_case_record_workbook(
            state, load_master_formulas().head(0), load_master_modulars().head(0)
        )
        restored, _, _, _ = import_case_record_workbook(BytesIO(payload))
        self.assertEqual(restored["assessment_iv_fluid_0"], "D5 1/2 NS")
        self.assertEqual(restored["assessment_iv_rate_0"], 100.0)

    def test_unknown_fluid_is_refused(self):
        state = {
            "case_record_label": "IV case",
            "assessment_iv_fluid_0": "Something this build does not know",
            "assessment_iv_rate_0": 100.0,
        }
        payload = export_case_record_workbook(
            state, load_master_formulas().head(0), load_master_modulars().head(0)
        )
        with self.assertRaises(ValueError):
            import_case_record_workbook(BytesIO(payload))

    def test_negative_rate_is_refused(self):
        state = {
            "case_record_label": "IV case",
            "assessment_iv_fluid_0": "D5W",
            "assessment_iv_rate_0": -10.0,
        }
        payload = export_case_record_workbook(
            state, load_master_formulas().head(0), load_master_modulars().head(0)
        )
        with self.assertRaises(ValueError):
            import_case_record_workbook(BytesIO(payload))
